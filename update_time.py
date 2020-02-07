from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from pathlib import Path
import time
import os
from xlrd import open_workbook
import argparse
from pynput.keyboard import Key, Controller
import warnings

startTime = time.time();
warnings.filterwarnings("ignore", category=DeprecationWarning)
keyboard = Controller()

parser = argparse.ArgumentParser(description="excel sheet time update")
parser.add_argument('firstname', type=str)
parser.add_argument('lastname', type=str)
parser.add_argument('course', type=str)
args = parser.parse_args()

print("Automating to USA swimming website...")
# startup webdriver
driver = webdriver.Chrome(os.getcwd() + "/chromedriver")
driver.set_window_size(1000, 1500)
driver.get("http://www.google.com")
driver.set_script_timeout(30)
# search usa swimming on google
searchBox = driver.find_element_by_name("q")
searchBox.send_keys("https://www.usaswimming.org")
searchBox.submit()
# navigate to usa swimming website
UsaSwimmingLink = driver.find_element(By.XPATH, '//*[@id="rso"]/div[1]/div/div/div/div/div[1]/a/h3')
UsaSwimmingLink.click()
# Navigate to time search
print("Navigating to time search...")
timeSearchTab = driver.find_element(By.XPATH, '//*[@id="main-collapsible-menu"]/ul/li[3]/a')
timeSearchTab.click()
timeSearchLink = driver.find_element(By.XPATH, '//*[@id="main-collapsible-menu"]/ul/li[3]/ul/li/div/div[3]/a/h4')
timeSearchLink.click()

CLUB_NAME = "Ames Cyclone Aquatics Club"
SCYevents = ["50 FR", "100 FR", "200 FR", "500 FR", "1000 FR", "1650 FR", "50 BK", "100 BK", "200 BK", "50 BR",
             "100 BR",
             "200 BR", "50 FL", "100 FL", "200 FL", "100 IM", "200 IM", "400 IM"]
LCMEvents = ["50 FR", "100 FR", "200 FR", "400 FR", "800 FR", "1500 FR", "50 BK", "100 BK", "200 BK", "50 BR", "100 BR",
             "200 BR", "50 FL", "100 FL", "200 FL", "100 IM", "200 IM", "400 IM"]
event_colums = ["F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W"]

courseInfo = ""
workbookPath = ""
if (args.course == 's'):
    workbookPath = os.getcwd() + "/data/Shortcourse.xlsx"
    wb = load_workbook(workbookPath)
    ws = wb.get_sheet_by_name("Shortcourse")
    events = SCYevents
    courseInfo = "Shortcourse"
    fileEnder = "Short"
elif (args.course == 'l'):
    workbookPath = os.getcwd() + "/data/Longcourse.xlsx"
    wb = load_workbook(workbookPath)
    ws = wb.get_sheet_by_name("Longcourse")
    events = LCMEvents
    courseInfo = "Longcourse"
    fileEnder = "Long"

# Start
nameIndex = 39
# Find number of people in sheet
numPeople = 0
findPeopleStart = nameIndex
while (ws["A" + str(findPeopleStart)].value != 'Template' and ws["A" + str(findPeopleStart)].value != None):
    numPeople = numPeople + 1
    findPeopleStart = findPeopleStart + 8
# flags for only updating one person and for setting the course type(long or short) only once.
oneFlag = False
courseTypeDone = False

if ((args.firstname + " " + args.lastname) != "full update"):
    oneFlag = True
    oneNameIndex = 39
    while (ws["A" + str(oneNameIndex)].value != 'Template' and ws["A" + str(oneNameIndex)].value != (
            args.firstname + " " + args.lastname)):
        oneNameIndex = oneNameIndex + 8
    nameIndex = oneNameIndex
    oneFirstName = args.firstname
    oneLastName = args.lastname
    ws['A' + str(nameIndex)].value = oneFirstName + " " + oneLastName
print("Number Of Athletes = " + str(numPeople))
time.sleep(0.2)
print("nameIndex = " + str(nameIndex))
time.sleep(0.2)
print("Course = " + courseInfo)
time.sleep(0.2)

print("-----------------------------------------------------------")
# iterate through each name on my excel sheet
while (ws["A" + str(nameIndex)].value != None and ws["A" + str(nameIndex)].value != 'Template'):
    name = ws["A" + str(nameIndex)].value
    nameList = name.split()
    if (oneFlag == False):
        firstName = nameList[0]
        lastName = nameList[1]
    else:
        firstName = oneFirstName
        lastName = oneLastName
    firstNameBox = driver.find_element(By.XPATH, '//*[@id="UsasTimeSearchIndividual_Index_Div_1FirstName"]')
    firstNameBox.clear()
    firstNameBox.send_keys(firstName)
    lastNameBox = driver.find_element(By.XPATH, '//*[@id="UsasTimeSearchIndividual_Index_Div_1LastName"]')
    lastNameBox.clear()
    lastNameBox.send_keys(lastName)

    if (courseTypeDone == False):
        if (args.course == "s"):
            lcmButton = driver.find_element_by_xpath(
                '//*[@id="UsasTimeSearchIndividual_Index_Div_1"]/div[1]/form/fieldset/div[4]/div/div[3]/span/span')
            lcmButton.click()
            keyboard.press('s')
            keyboard.release('s')
            keyboard.press('c')
            keyboard.release('c')
            keyboard.press('y')
            keyboard.release('y')
            time.sleep(0.2)
        else:
            lcmButton = driver.find_element_by_xpath(
                '//*[@id="UsasTimeSearchIndividual_Index_Div_1"]/div[1]/form/fieldset/div[4]/div/div[3]/span/span')
            lcmButton.click()
            keyboard.press('l')
            keyboard.release('l')
            time.sleep(0.2)
        courseTypeDone = True
    findTimesButton = driver.find_element(By.XPATH, '//*[@id="UsasTimeSearchIndividual_Index_Div_1-saveButton"]')
    findTimesButton.click()
    time.sleep(1)
    # if they are the only name, download excel file
    excelLinkExists = True
    try:
        driver.find_element(By.XPATH, '//*[@id="UsasTimeSearchIndividual_TimeResults_Grid-1_exportExcel"]')
    except:
        excelLinkExists = False

    nameTableExists = False
    while (nameTableExists == False):
        try:
            nameTable = driver.find_element(By.XPATH,
                                            '//*[@id="UsasTimeSearchIndividual_PersonSearchResults_Grid-1"]/table/tbody')
            nameTableExists = True
        except:
            time.sleep(1)
    # otherwise, look through rows
    if (excelLinkExists == True):
        one = 1
    else:
        found = False
        numTries = 0
        while (numTries < 6):
            numTries = numTries + 1
            # if the name is not found, the name was typed wrong and their times will not update.
            # It will move on to the next name
            rows = nameTable.find_elements(By.TAG_NAME, "tr")
            numRows = len(rows)
            for i in range(numRows):
                clubNameTexts = rows[i].find_elements(By.TAG_NAME, "td")
                clubName = clubNameTexts[1].text
                if (clubName == CLUB_NAME):
                    textNameLink = clubNameTexts[0].find_elements(By.TAG_NAME, "a")
                    textNameLink[0].click()
                    found = True
                    break
            if (found == True):
                break
            else:
                nextListButton = driver.find_element(By.XPATH,
                                                     '//*[@id="UsasTimeSearchIndividual_PersonSearchResults_Grid-1-UsasGridPager-pgNext"]')
                nextListButton.click()

            # if name is typed incorrectly in sheet
        if (found == False):
            nameIndex = nameIndex + 8
            print(firstName + " " + lastName + " NOT PROCESSED")
            if (oneFlag == False):
                continue
            else:
                break
    # if the name is found but no table of events shows up, then they have not swam that course
    eventTableExists = False
    while (eventTableExists == False):
        try:
            eventTable = driver.find_element_by_xpath('//*[@id="UsasTimeSearchIndividual_TimeResults_Grid-1"]')
            eventTableExists = True
        except:
            time.sleep(1)

    if (eventTableExists == False):
        nameIndex = nameIndex + 8
        print(firstName + " " + lastName + " Has no events for this course")
        if (oneFlag == False):
            continue
        else:
            break
    time.sleep(0.3)
    excelDownloadLink = driver.find_element_by_id("UsasTimeSearchIndividual_TimeResults_Grid-1_exportExcel")
    excelDownloadLink.click()
    fileNameHeader = driver.find_element(By.XPATH, '//*[@id="UsasTimeSearchIndividual_Index_Div_1-Results"]/h3')
    fileNameText = fileNameHeader.text
    fileNameWords = fileNameText.split()
    print(fileNameWords)
    finalFileName = ""
    for j in range(4):
        current = fileNameWords[j]
        current = str(current).lower().capitalize()
        finalFileName = finalFileName + current
        if (j < 3):
            finalFileName = finalFileName + " "
    # wait for file to download
    print(finalFileName)
    downloadPath = str(Path.home()) + "/Downloads"
    print(downloadPath + "/" + finalFileName + ".xlsx")
    while (os.path.exists(downloadPath + "/" + finalFileName + ".xlsx") == False):
        time.sleep(0.5)
    os.rename(downloadPath + "/" + finalFileName + ".xlsx",
              os.getcwd() + "/data/Personal_Sheets/" + firstName + lastName + ".xlsx")
    # open up swimmers times sheet
    swimmerWb = load_workbook(os.getcwd() + "/data/Personal_Sheets/" + firstName + lastName + ".xlsx")
    timeSheet = swimmerWb.active
    stupidWb = open_workbook(os.getcwd() + "/data/Personal_Sheets/" + firstName + lastName + ".xlsx")
    stupidSheet = stupidWb.sheet_by_index(0)
    betterIndex = nameIndex

    # iterate through swimmers time sheet and add best times from each event to main sheet

    for event in range(18):
        flag = False
        index = 1
        current = events[event]
        while (timeSheet["A" + str(index)].value != None):
            if (timeSheet["A" + str(index)].value == current):
                swimmerSheetCell = stupidSheet.cell(index - 1, 1)
                strFlag = False
                if (isinstance(swimmerSheetCell.value, str) == True):
                    strFlag = True
                    swimmerSheetCell.value = swimmerSheetCell.value[:-1]

                swimmerWb.save(
                    os.getcwd() + "/data/Personal_Sheets/" + firstName + lastName + ".xlsx")

                c = ws[event_colums[event] + str(betterIndex)]
                if (strFlag == True):
                    c.value = "00:" + swimmerSheetCell.value
                else:
                    c.value = swimmerSheetCell.value - 33.5
                # c.number_format = 'm:ss.00'
                c.font = Font(size=12, bold=True)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                flag = True
                break
            index = index + 1
        if (flag == False):
            d = ws[event_colums[event] + str(betterIndex)]
            d.value = 0
            d.number_format = 'm:ss.00'
            d.font = Font(size=12, bold=True)
            d.alignment = Alignment(horizontal='center', vertical='center')
            d.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    print(firstName + " " + lastName + " UPDATED")
    if (oneFlag == True):
        break
    nameIndex = nameIndex + 8

# backup copy
# wb.save("/Users/nickmatthews/OneDrive - Iowa State University/Final-Pre-Senior-Times-" + fileEnder + ".xlsx")

# Fix -1 problem
fixWb = open_workbook(workbookPath)
fixWs = fixWb.sheet_by_index(0)

startZeroFix = 38
testCount = 0
fixCount = 0
for people in range(numPeople):
    for luna in range(18):
        lunaCell = fixWs.cell(startZeroFix, luna + 5)
        if (isinstance(lunaCell.value, float) == True):
            if (int(lunaCell.value) < 0):
                fixCount = fixCount + 1
                testCount = testCount + 1
                fixCell = ws[event_colums[luna] + str(startZeroFix + 1)]
                fixCell.value = 0
                fixCell.number_format = 'm:ss.00'
                fixCell.font = Font(size=12, bold=True)
                fixCell.alignment = Alignment(horizontal='center', vertical='center')
                fixCell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    startZeroFix = startZeroFix + 8
print("Number of Cells Fixed = " + str(fixCount))
print("Saving Sheets...")

# wb.save("/Users/nickmatthews/OneDrive - Iowa State University/Final-Pre-Senior-Times-" + fileEnder + ".xlsx")
wb.save(os.getcwd() + "/data/" + fileEnder + "course.xlsx")

timestr = time.strftime("%Y%m%d-%H%M%S")
wb.save(os.getcwd() + "/data/Update_Log/log_" + timestr + ".xlsx")
print("----------------------------------------------------------")
print("Update Complete")
wb.close()
driver.close()

endTime = time.time()
excecutionTime = endTime - startTime
minutes = int(excecutionTime / 60)
seconds = int(excecutionTime % 60)
print("Total Time of Excecution = " + str(minutes) + ":" + str(seconds))
